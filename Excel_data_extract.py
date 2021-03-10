#Script to extract Straps connections with values from Ajmal's sheet
#By Bhanu Pratap  Singh
# importing openpyxl module
import openpyxl as xl;

# opening the source excel file
filename = "C:\\Users\\singhb\\OneDrive - Intel Corporation\\Desktop\\Intel\\Formal Verification\\Straps\\rwc_10x7_straps_spec.xlsx"
wb1 = xl.load_workbook(filename)
ws1 = wb1.worksheets[43]

# opening the destination excel file
filename1 = "C:\\Users\\singhb\\OneDrive - Intel Corporation\\Desktop\\Intel\\Formal Verification\\Straps\\straps_extract.xlsx"
wb2 = xl.load_workbook(filename1)
ws2 = wb2.worksheets[0]

# calculate total number of rows and
# columns in source excel file
mr = ws1.max_row
mc = ws1.max_column

# copying the cell values from source
# excel file to destination excel file
# defining some variables
dst_col_count = 5
row_count_col4 = 1
row_count_col2_3 = 1
for j in range(1, mc + 1):
    for i in range(2, mr + 1):
        if j == dst_col_count:
            #print('inside',i, j)
            # reading cell value from source excel file for column 5 onwards and every rows
            v = ws1.cell(row=i, column=j)
            # writing the read value to destination excel file in column of values
            ws2.cell(row=row_count_col2_3, column=2).value = v.value

            d = ws1.cell(row=1, column=j)
            # writing the read value to destination excel file in column of dst
            ws2.cell(row=row_count_col2_3, column=3).value = d.value
            row_count_col2_3+=1
        else:
            break
        # reading cell value from source excel file for column 1 and each row
        s = ws1.cell(row=i, column=1)
        # writing the read value to destination excel file in column of straps
        ws2.cell(row=row_count_col4, column=4).value = s.value
        row_count_col4 += 1

    if j == dst_col_count:
        dst_col_count += 1

# saving the destination excel file
wb2.save(filename1)
