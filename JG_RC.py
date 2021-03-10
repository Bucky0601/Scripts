# importing openpyxl module
import openpyxl as xl;


# opening the source excel file
filename = "C:\\Users\\singhb\\OneDrive - Intel Corporation\\Desktop\\Intel\\Formal Verification\\Straps\\test_for_tcl_file.xlsx"
wb1 = xl.load_workbook(filename)
ws1 = wb1.worksheets[0]


# creating the destination tcl file
f = open("C:\\Users\\singhb\\OneDrive - Intel Corporation\\Desktop\\Intel\\Formal Verification\\Straps\\tcl_straps.tcl","w+")

# calculate total number of rows and
# columns in source excel file
mr = ws1.max_row
mc = ws1.max_column

for i in range(2, mr + 1):
    d = ws1.cell(row=i, column=5)
    # reading cell value from source excel file for column 1 and each row
    v = ws1.cell(row=i, column=6)
    # writing the actual data of excel like dst block and dst signals for RC target purpose
    f.write("check_conn -reverse -target {%s.%s} -relative_depth 5 -complexity simple -stop_on_target_boundary on  -silent -save_as /nfs/site/home/singhb/strap.csv -append -force \n" % (d.value,v.value))

#closing file here
f.close()